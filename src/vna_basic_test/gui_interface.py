from multi_thread import Progress
from visa_scpi import Vna_measure

import time
from time import gmtime, strftime

from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog


from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg, NavigationToolbar2Tk)
from matplotlib.backend_bases import key_press_handler
from matplotlib.figure import Figure
import matplotlib.pyplot as plt

import numpy as np

from openpyxl import *


# variables
reference = False
saveRef = False 

        
class Frame_examples_program():
    
    def __init__(self):
        self.window = tk.Tk()
        #self.window.geometry('600x600')

        # opening the existing excel file & create the sheet object
        self.wb = Workbook()
        self.sheet = self.wb.active

        self.vna = Vna_measure()
     
        self.window.title("TEST GUI - V.1.0")
        self.create_widgets()        


    def configure(self, event):
        # read screen width & height
        w, h = event.width, event.height


    def focus(self, event):
        self.create_plot()
        

    def update_screen(self):
        self.clock.config(text="DATE: " + strftime("%d %b %Y %H:%M:%S", gmtime()))        
        # update every 1000ms
        self.clock.after(1000, self.update_screen)
        

    def create_buttons(self, parent, a, b, c):
        button1 = ttk.Button(parent, text="do task " + a)
        button1.grid(row=1, column=1)
        button2 = ttk.Button(parent, text="do task " + b)
        button2.grid(row=2, column=1)
        button3 = ttk.Button(parent, text="do task " + c)
        button3.grid(row=3, column=1)
        return (button1, button2, button3)


    def panel_led(self, color):
        self.circle_canvas.create_oval(10, 10, 40, 40, width=0, fill=color)
	
	
    def show_info(self):
        messagebox.showinfo(title = 'About Me!', message = 'joel.daricou@cern.ch 2018')


    def save_file(self):    
        file_name = self.serial_name_field.get() + '_' + self.serial_number_field.get() + '_' + self.details_field.get()    
        file_position = filedialog.asksaveasfilename(initialdir = "/",initialfile=file_name, \
        title = "Select file",filetypes = (("Microsoft Excel Worksheet","*.xlsx"),("all files","*.*")), defaultextension = ' ')
        self.wb.save()


    def open_file(self):
        #self.file_position = filedialog.askopenfile().name
        return
        

    def close_file(self):
        exit = messagebox.askyesno(title = 'Quit?', message = 'Are you sure?')
        if exit > 0:
            self.window.destroy()


    def save_ref(self):
        global saveRef
        global reference
        # invert the value
        reference = not reference
        if reference == True:
            self.save_ref.config(text='Remove Ref.')
            saveRef = True
        else:
            self.save_ref.config(text='Save Ref.')


    def start_test(self):
        self.display_info.config(text='START TEST')
        self.panel_led('lime')
        self.create_plot()
        

    def set_axis_name(self):
        # set axis names
        self.plot0.set_title('S21 - delay')
        self.plot0.set_xlabel('freq')
        self.plot0.set_ylabel('dB')

        self.plot1.set_title('S21 - dB')
        self.plot1.set_xlabel('freq')
        self.plot1.set_ylabel('dB')

        self.plot2.set_title('S11 - SWR')
        self.plot2.set_xlabel('freq')
        self.plot2.set_ylabel('dB')

        self.plot3.set_title('S22 - SWR')
        self.plot3.set_xlabel('freq')
        self.plot3.set_ylabel('dB')

        self.plot4.set_title('S11 - TDR')
        self.plot4.set_xlabel('delay')
        self.plot4.set_ylabel('dB')

                       
    def create_plot(self):              
        # masure vna
        xValue0, yValue0 = self.vna.read_measure(0)
        xValue1, yValue1 = self.vna.read_measure(1)
        xValue2, yValue2 = self.vna.read_measure(2)
        xValue3, yValue3 = self.vna.read_measure(3)
        xValue4, yValue4 = self.vna.read_measure(4)

        global saveRef
        if saveRef == True:
            # set reference data on plot
            self.xRef0, self.yRef0 = xValue0, yValue0
            self.xRef1, self.yRef1 = xValue1, yValue1
            self.xRef2, self.yRef2 = xValue2, yValue2
            self.xRef3, self.yRef3 = xValue3, yValue3
            self.xRef4, self.yRef4 = xValue4, yValue4
            saveRef = False           
          
        # clean plot line
        self.plot0.cla()
        self.plot1.cla()
        self.plot2.cla()
        self.plot3.cla()
        self.plot4.cla()

        # set axis names
        self.set_axis_name()
      
        # set data on plot
        self.plot0.plot(xValue0, yValue0)        
        self.plot1.plot(xValue1, yValue1)
        self.plot2.plot(xValue2, yValue2)
        self.plot3.plot(xValue3, yValue3)
        self.plot4.plot(xValue4, yValue4)

        global reference
        if reference == True:     
            self.plot0.plot(self.xRef0, self.yRef0)        
            self.plot1.plot(self.xRef1, self.yRef1)
            self.plot2.plot(self.xRef2, self.yRef2)
            self.plot3.plot(self.xRef3, self.yRef3)
            self.plot4.plot(self.xRef4, self.yRef4)

        # update plot    
        self.canvas.draw()
        
        # - - - - - - - - - - - - - - - - - - - - -
        # Create sheet
        """
        # resize the width of columns in excel spreadsheet
        self.sheet.column_dimensions['A'].width = 30
        self.sheet.column_dimensions['B'].width = 30
        """

        # write given data to an excel spreadsheet at particular location
        file_name = self.name_field.get() + '_' +  self.serial_name_field.get() + '_' + \
        self.serial_number_field.get() + '_' + self.details_field.get()

        self.sheet.cell(row=1, column=1).value = file_name

        self.sheet.cell(row=1, column=2).value =strftime("%d %b %Y %H:%M:%S", gmtime())

        self.sheet.cell(row=2, column=1).value = 'x'
        self.sheet.cell(row=2, column=2).value = 'y'

        self.sheet.cell(row=2, column=4).value = 'x'
        self.sheet.cell(row=2, column=5).value = 'y'

        self.sheet.cell(row=2, column=7).value = 'x'
        self.sheet.cell(row=2, column=8).value = 'y'

        self.sheet.cell(row=2, column=10).value = 'x'
        self.sheet.cell(row=2, column=11).value = 'y'

        self.sheet.cell(row=2, column=13).value = 'x'
        self.sheet.cell(row=2, column=14).value = 'y'
        
        for i in range(0, len(xValue0), 1):
            self.sheet.cell(row=i + 3, column=1).value = xValue0[i]
            self.sheet.cell(row=i + 3, column=2).value = yValue0[i]

            self.sheet.cell(row=i + 3, column=4).value = xValue1[i]
            self.sheet.cell(row=i + 3, column=5).value = yValue1[i]

            self.sheet.cell(row=i + 3, column=7).value = xValue2[i]
            self.sheet.cell(row=i + 3, column=8).value = yValue2[i]

            self.sheet.cell(row=i + 3, column=10).value = xValue3[i]
            self.sheet.cell(row=i + 3, column=11).value = yValue3[i]

            self.sheet.cell(row=i + 3, column=13).value = xValue4[i]
            self.sheet.cell(row=i + 3, column=14).value = yValue4[i]
            
        self.display_info.config(text='TEST DONE')
        self.save_file()

            
    def create_widgets(self):
        
        # - - - - - - - - - - - - - - - - - - - - -
        # MenuBar
        myMenuBar = Menu (self.window)

        myFileMenu = Menu (myMenuBar , tearoff = 0)
        myFileMenu.add_command(label = "Exit", command = self.close_file)
        myFileMenu.add_command(label = "Open", command = self.open_file)
        myFileMenu.add_command(label = "Save as", command = self.save_file)
        myMenuBar.add_cascade(label = "File", menu = myFileMenu)

        myFileMenu = Menu (myMenuBar , tearoff = 0)
        myFileMenu.add_command(label = "Info", command = self.show_info)
        myMenuBar.add_cascade(label = "Help", menu = myFileMenu)

        self.window.config(menu = myMenuBar)

        # Create some room around all the internal frames
        self.window['padx'] = 10
        self.window['pady'] = 10      

        # - - - - - - - - - - - - - - - - - - - - -
        # Title
        instrument_name, instrument_address = self.vna.instrument_info()
        str_instrument_info = instrument_name + "\n" + instrument_address
        
        labeled_frame_label = ttk.Label(self.window, text=str_instrument_info)
        labeled_frame_label.grid(row=0, column=0, sticky=W, padx=10, pady=5)

        # - - - - - - - - - - - - - - - - - - - - -
        # User data
        frame = ttk.LabelFrame(self.window, text="USER DATA", relief=tk.RIDGE)
        frame.grid(row=1, column=1, sticky = tk.E + tk.W + tk.N + tk.S, padx=10, pady=10)
       
        name = Label(frame, text='Name (User):')
        name.grid(row=0, column=0, sticky = W)
        self.name_field = Entry(frame)
        self.name_field.grid(row=0, column=1, padx=5, pady = 5)

        serial_name = Label(frame, text='Serial name:')
        serial_name.grid(row=1, column=0, sticky = W)
        self.serial_name_field = Entry(frame)
        self.serial_name_field.grid(row=1, column=1, padx=5, pady = 5)

        serial_number = Label(frame, text='Serial num.:')
        serial_number.grid(row=2, column=0, sticky = W)
        self.serial_number_field = Entry(frame)
        self.serial_number_field.grid(row=2, column=1, padx=5, pady = 5)

        details = Label(frame, text='Add details:')
        details.grid(row=3, column=0, sticky = W)
        self.details_field = Entry(frame)
        self.details_field.grid(row=4, column=0, sticky = E + W, columnspan=2, padx=5, pady = 5)
        
        self.save_ref = Button(frame, text='Save ref.', fg='Black', command= self.save_ref)
        self.save_ref.grid(row=10, column=0, columnspan=2, sticky = W, padx=20, pady = 5)
        
        submit = Button(frame, text='START MEASURE', fg='Black', command= self.start_test)
        submit.grid(row=10, column=1, columnspan=2, sticky = E, padx=20, pady = 5)

        # whenever the enter key is pressed then call the focus1 function
        self.window.bind('<Return>', self.focus)

        self.display_info = ttk.Label(frame, text= 'PRESS TO START')
        self.display_info.grid(row=12, column=0, sticky = tk.E + tk.W + tk.N + tk.S, padx=5, pady=20)

        self.circle_canvas = Canvas(frame, width=40, height=40)
        self.circle_canvas.grid(row=12, column=1, sticky = tk.E + tk.W + tk.N + tk.S, padx=0, pady=5)

        self.clock = Label(frame)
        self.clock.grid(row=13, column=0, sticky = tk.E + tk.W + tk.N + tk.S, padx=5, pady=5)

        # - - - - - - - - - - - - - - - - - - - - -
        # test thread
        parent = ttk.LabelFrame(frame, text="NOTEBOOK", relief=tk.RIDGE)
        parent.grid(row=21, column=0, sticky = tk.E + tk.W + tk.N + tk.S, padx=10, pady=10)
        
        prog_bar = Progress(parent, row=0, column=0, columnspan=2)
        # Button 1
        start_button = ttk.Button(parent, text="start",
                                  command=prog_bar.pb_start)
        start_button.grid(row=1, column=0)
        # Button 2
        stop_button = ttk.Button(parent, text="stop",
                                 command=prog_bar.pb_stop)
        stop_button.grid(row=1, column=1)
        # Button 3
        complete_button = ttk.Button(parent, text="complete",
                                     command=prog_bar.pb_complete)
        complete_button.grid(row=2, column=0)
        # Button 4
        clear_button = ttk.Button(parent, text="clear",
                                  command=prog_bar.pb_clear)
        clear_button.grid(row=2, column=1)

        """
        # - - - - - - - - - - - - - - - - - - - - -
        # Setup
        frame2 = ttk.LabelFrame(frame, text="NOTEBOOK", relief=tk.RIDGE)
        frame2.grid(row=21, column=0, sticky = tk.E + tk.W + tk.N + tk.S, padx=20, pady=20)
       
        frame3 = ttk.Notebook(frame2)
        frame3.grid(row=1, column=0, sticky=tk.E + tk.W + tk.N + tk.S, padx=0, pady=0)
               
        tab1 = tk.Frame(frame2)
        tab2 = tk.Frame(frame2)

        frame3.add(tab1, text="TEST", compound=tk.TOP)
        frame3.add(tab2, text="SETUP", compound=tk.TOP)
        
        self.create_buttons(tab1, "J", "K", "L")
        self.create_buttons(tab2, "M", "N", "O")
        """

        # - - - - - - - - - - - - - - - - - - - - -
        # Plot setup
        plt.style.use('bmh')
        
        self.fig = Figure()       
        #self.fig.set_size_inches(22, 10)

        self.fig = plt.gcf()
        DPI = self.fig.get_dpi()
        self.fig.set_size_inches(1280.0/float(DPI), 720.0/float(DPI))
        
        # - - - - - - - - - - - - - - - - - - - - -
        # Plot 0
        self.plot0 = self.fig.add_subplot(231)
                                               
        # - - - - - - - - - - - - - - - - - - - - -
        # Plot 1       
        self.plot1 = self.fig.add_subplot(232)

        # - - - - - - - - - - - - - - - - - - - - -
        # Plot 2       
        self.plot2 = self.fig.add_subplot(233)

        # - - - - - - - - - - - - - - - - - - - - -
        # Plot 3       
        self.plot3 = self.fig.add_subplot(234)

        # - - - - - - - - - - - - - - - - - - - - -
        # Plot 4       
        self.plot4 = self.fig.add_subplot(235)

        # - - - - - - - - - - - - - - - - - - - - -
        # set axis names
        self.set_axis_name()

        # - - - - - - - - - - - - - - - - - - - - -
        # autoadapt
        plt.tight_layout()

        # - - - - - - - - - - - - - - - - - - - - -
        # Draw        
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.window)
        self.canvas.draw()
        self.canvas.get_tk_widget().grid(row=1, column=0, sticky=tk.E + tk.W + tk.N + tk.S, padx=10, pady=10)

        # - - - - - - - - - - - - - - - - - - - - -
        # event screen resize
        self.window.bind("<Configure>", self.configure)
        self.update_screen()

